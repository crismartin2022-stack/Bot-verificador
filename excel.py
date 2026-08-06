"""Lectura del Excel del admin y cruce contra lo verificado en el chat."""
from __future__ import annotations

import re
import logging
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

import config
import utils

log = logging.getLogger("excel")

CLAVES_NOMBRE = ("nombre", "apellido", "titular", "cliente", "beneficiario",
                 "destinatario", "socio", "alumno", "persona", "paciente")
CLAVES_MONTO = ("monto", "importe", "valor", "pago", "total", "abono", "deposito", "depósito")
CLAVES_FECHA = ("fecha", "dia", "día")
CLAVES_DOC = ("cuit", "cuil", "dni", "documento", "doc")
CLAVES_HORA = ("hora",)
ENCABEZADOS_RUIDO = {"NOMBRE", "MONTO", "IMPORTE", "TITULAR", "CLIENTE", "FECHA", "TOTAL",
                     "APELLIDO", "NOMBRE Y APELLIDO", "BENEFICIARIO", "DESTINATARIO"}

# Peso de cada palabra por campo. Más alto = más confiable.
# Evita que «TRF O DEPOSITO» gane sobre «MONTO», o «TITULAR DE LA CTA»
# (la empresa que cobra) sobre «Remitente» (el cliente que paga).
PESOS = {
    "nombre": [("nombre y apellido", 10), ("remitente", 10), ("pagador", 10),
               ("nombre", 9), ("apellido", 8), ("cliente", 8), ("beneficiario", 7),
               ("titular", 6), ("destinatario", 5), ("socio", 5), ("alumno", 5),
               ("persona", 5), ("paciente", 5)],
    "monto": [("monto", 10), ("importe", 10), ("valor", 7), ("total", 6),
              ("abono", 5), ("pago", 5), ("deposito", 3)],
    "doc": [("cuit", 10), ("cuil", 10), ("dni", 10), ("documento", 9),
            ("nro doc", 9), ("doc", 7), ("cvu", 4), ("cbu", 4)],
    "fecha": [("fecha ticket", 10), ("fecha op", 10), ("fecha operacion", 10),
              ("fecha", 7), ("dia", 4)],
    "hora": [("hora ticket", 10), ("hora", 9)],
    "oper": [("nro operacion", 10), ("n operacion", 10), ("operacion", 9),
             ("codigo", 8), ("referencia", 8), ("comprobante", 6)],
    "notas": [("notas", 10), ("observaciones", 10), ("observacion", 9), ("detalle", 6)],
}


def _peso(titulo: str, campo: str) -> float:
    t = utils.normalizar(titulo).lower()
    if not t:
        return 0.0
    mejor = 0.0
    for clave, peso in PESOS[campo]:
        if t == clave:
            calidad = 1.0
        elif t.startswith(clave) or t.endswith(clave):
            calidad = 0.9
        elif clave in t:
            calidad = 0.7
        else:
            continue
        mejor = max(mejor, peso * calidad)
    return mejor


def _detectar_columnas(hoja, max_filas: int = 15) -> tuple[int, dict]:
    """Devuelve (fila_encabezado, {campo: nº de columna}).

    Se elige la mejor columna para cada campo, no la primera que coincida.
    """
    for fila in range(1, min(max_filas, hoja.max_row or 1) + 1):
        opciones = []
        for col in range(1, min(hoja.max_column or 1, 40) + 1):
            val = hoja.cell(row=fila, column=col).value
            if not isinstance(val, str):
                continue
            for campo in PESOS:
                p = _peso(val, campo)
                if p > 0:
                    opciones.append((p, campo, col))

        mapa: dict[str, int] = {}
        usadas: set[int] = set()
        for _, campo, col in sorted(opciones, key=lambda x: -x[0]):
            if campo in mapa or col in usadas:
                continue
            mapa[campo] = col
            usadas.add(col)

        if "nombre" in mapa and "monto" in mapa:
            return fila, mapa
    # Sin encabezados reconocibles: asumimos A=nombre, B=monto
    return 0, {"nombre": 1, "monto": 2}


def leer_excel(path: str | Path, hojas: list[str] | None = None) -> tuple[list[dict], str]:
    """Lee las hojas del libro (todas, o solo las de `hojas`).

    Devuelve (filas, cómo se interpretó).
    """
    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    filas: list[dict] = []
    detalles: list[str] = []
    ignoradas: list[str] = []

    pedidas = [utils.normalizar(h) for h in (hojas or []) if h]
    seleccionadas = wb.worksheets
    if pedidas:
        seleccionadas = [h for h in wb.worksheets if utils.normalizar(h.title) in pedidas]
        if not seleccionadas:
            disponibles = ", ".join(h.title for h in wb.worksheets)
            wb.close()
            return [], f"no encontré esa(s) hoja(s). Disponibles: {disponibles}"

    for hoja in seleccionadas:
        fila_enc, cols = _detectar_columnas(hoja)
        if not fila_enc and hoja.max_row and hoja.max_row > 2:
            # Sin encabezado reconocible: solo se asume A/B si es la única hoja
            if len(seleccionadas) > 1:
                ignoradas.append(hoja.title)
                continue

        n_antes = len(filas)
        inicio = fila_enc + 1 if fila_enc else 1

        for i, fila in enumerate(hoja.iter_rows(min_row=inicio, values_only=True), start=inicio):
            def celda(clave):
                idx = cols.get(clave)
                if not idx or idx > len(fila):
                    return None
                return fila[idx - 1]

            nombre_raw = celda("nombre")
            monto_raw = celda("monto")
            if nombre_raw is None and monto_raw is None:
                continue
            nombre = str(nombre_raw).strip() if nombre_raw is not None else ""
            if utils.normalizar(nombre) in ENCABEZADOS_RUIDO:
                continue  # encabezado repetido en medio de la hoja
            monto = utils.parse_monto(monto_raw)
            if not utils.tokens_nombre(nombre) and monto is None:
                continue
            fecha = celda("fecha")
            if isinstance(fecha, datetime):
                fecha = fecha.date().isoformat()
            hora = celda("hora")
            if hasattr(hora, "strftime"):
                hora = hora.strftime("%H:%M")

            # Código de operación: columna propia o dentro de las notas
            oper = str(celda("oper") or "").strip()
            notas = str(celda("notas") or "")
            doc = utils.clave_doc(celda("doc"))
            if notas:
                if not oper:
                    m_op = re.search(r"(?i)c[oó]digo[^:]*:\s*([A-Z0-9-]{6,})", notas)
                    if m_op:
                        oper = m_op.group(1)
                if not doc:
                    m_cvu = re.search(r"(?i)cvu\s*remitente\s*:?\s*(\d{10,22})", notas)
                    if m_cvu:
                        doc = utils.clave_doc(m_cvu.group(1))

            filas.append({
                "hoja": hoja.title,
                "fila": i,
                "nombre": nombre,
                "monto": monto,
                "doc": doc,
                "oper": utils.normalizar(oper),
                "fecha": str(fecha) if fecha else "",
                "hora": str(hora) if hora else "",
            })

        if len(filas) > n_antes:
            detalles.append(f"«{hoja.title}» {len(filas) - n_antes} filas "
                            f"({get_column_letter(cols['nombre'])}/{get_column_letter(cols['monto'])})")
    wb.close()

    desc = f"{len(filas)} filas de {len(detalles)} hoja(s): " + "; ".join(detalles)
    if ignoradas:
        desc += " · sin encabezado reconocible: " + ", ".join(ignoradas)
    return filas, desc


# ------------------------------------------------------------------- cruce
def _datos_item(it: dict) -> tuple[str, float | None]:
    nombre = it.get("nombre_img") or it.get("nombre_pie") or ""
    monto = it.get("monto_img")
    if monto is None:
        monto = it.get("monto_pie")
    return nombre, monto


def _perfil(it: dict) -> dict:
    """Todos los datos comparables de un comprobante del chat.

    Se reinterpreta el pie con la lógica actual, así los comprobantes
    verificados con versiones viejas también aprovechan el DNI.
    """
    if it.get("_perfil"):
        return it["_perfil"]

    nombres = [it.get("nombre_img"), it.get("nombre_pie"),
               it.get("nombre_origen"), it.get("nombre_destino")]
    docs = [it.get("doc_pie"), it.get("cuit_origen"), it.get("cuit_destino"),
            it.get("cvu_destino")]
    montos = [it.get("monto_img"), it.get("monto_pie")]

    pie = it.get("pie") or ""
    if pie:
        n_pie, m_pie, d_pie = utils.parse_pie(pie)
        nombres.append(n_pie)
        docs.append(d_pie)
        montos.append(m_pie)

    perfil = {
        "opers": {utils.normalizar(o) for o in (it.get("nro_operacion"),) if o and len(str(o)) >= 6},
        "nombres": [n for n in nombres if n and utils.tokens_nombre(n)],
        "docs": {utils.clave_doc(d) for d in docs if utils.clave_doc(d)},
        "montos": [m for m in montos if m is not None],
        "fecha": (it.get("fecha_comp") or "")[:10],
        "fecha_msg": it.get("fecha_msg") or "",
    }
    it["_perfil"] = perfil
    return perfil


def _fecha_item(perfil: dict) -> str:
    """ISO yyyy-mm-dd de la operación, o del mensaje si no hay otra."""
    if perfil["fecha"]:
        return perfil["fecha"]
    m = re.match(r"(\d{2})/(\d{2})/(\d{4})", perfil["fecha_msg"])
    return f"{m.group(3)}-{m.group(2)}-{m.group(1)}" if m else ""


def _puntaje(perfil: dict, fila: dict) -> float:
    """-1 = no aparea. Si aparea, cuanto más alto, mejor candidato."""
    # 0) El número de operación es único: si coincide, es la misma transferencia
    if fila.get("oper") and fila["oper"] in perfil["opers"]:
        return 10.0

    # 1) El monto es el ancla: tiene que coincidir con alguno de los del chat
    if fila["monto"] is None or not perfil["montos"]:
        return -1.0
    if not any(utils.montos_iguales(m, fila["monto"], config.TOLERANCIA_MONTO)
               for m in perfil["montos"]):
        return -1.0

    # 2) Documento o nombre
    doc_fila = utils.clave_doc(fila.get("doc"))
    doc_ok = bool(doc_fila) and any(utils.docs_iguales(d, doc_fila) for d in perfil["docs"])
    sim = max([utils.similitud_nombres(n, fila["nombre"]) for n in perfil["nombres"]] or [0.0])

    if not doc_ok and sim < config.UMBRAL_NOMBRE:
        return -1.0

    puntaje = 2.0 if doc_ok else 0.0
    puntaje += sim

    # 3) Fecha y hora: solo desempatan, nunca descartan
    f_item = _fecha_item(perfil)
    if f_item and fila.get("fecha") and f_item == str(fila["fecha"])[:10]:
        puntaje += 0.5
    return puntaje


def cruzar(items_chat: list[dict], filas_excel: list[dict]) -> dict:
    """Clasifica en coinciden / duplicados / solo_chat / solo_excel.

    Aparea por monto + (documento o nombre). La fecha desempata.
    """
    candidatos = [
        it for it in items_chat
        if it.get("estado") not in ("no_comprobante", "sin_pie")
    ]

    # Duplicados: marcados durante la verificación, o repetidos dentro del lote
    # (incluye el mismo comprobante mandado a DOS grupos distintos).
    duplicados = [it for it in candidatos if it.get("estado") == "duplicado"]
    ids_dup = {id(it) for it in duplicados}
    vistos: dict[str, dict] = {}
    for it in candidatos:
        if it.get("estado") == "duplicado":
            continue
        claves = [k for k in (it.get("huella"), it.get("hash_archivo")) if k]
        if not claves:
            continue
        previo = next((vistos[k] for k in claves if k in vistos), None)
        if previo is not None:
            copia = dict(it)
            copia["dup_de"] = previo.get("msg_id")
            copia["dup_detalle"] = (f"ya contado en {previo.get('chat') or previo.get('chat_id')} "
                                    f"({previo.get('fecha_msg', '')})")
            duplicados.append(copia)
            ids_dup.add(id(it))       # el original queda fuera de los faltantes
        else:
            for k in claves:
                vistos[k] = it

    disponibles = [it for it in candidatos if id(it) not in ids_dup]
    perfiles = [_perfil(it) for it in disponibles]

    # Índice por monto: evita comparar cada fila contra todos los comprobantes
    indice: dict[float, list[int]] = {}
    for j, p in enumerate(perfiles):
        for m in p["montos"]:
            indice.setdefault(round(m, 2), []).append(j)

    usados: set[int] = set()
    coinciden, solo_excel = [], []

    por_oper: dict[str, int] = {}
    for j, p in enumerate(perfiles):
        for o in p["opers"]:
            por_oper.setdefault(o, j)

    for fila in filas_excel:
        mejor, mejor_p = None, 0.0
        j_op = por_oper.get(fila.get("oper") or "\x00")
        if j_op is not None and j_op not in usados:
            usados.add(j_op)
            coinciden.append({"excel": fila, "chat": disponibles[j_op],
                              "similitud": 1.0, "por_doc": True})
            continue
        if fila["monto"] is not None:
            cercanos = set(indice.get(round(fila["monto"], 2), []))
            if config.TOLERANCIA_MONTO:
                for m, js in indice.items():
                    if abs(m - fila["monto"]) <= config.TOLERANCIA_MONTO:
                        cercanos.update(js)
            for j in cercanos:
                if j in usados:
                    continue
                p = _puntaje(perfiles[j], fila)
                if p > mejor_p:
                    mejor, mejor_p = j, p
        if mejor is not None:
            usados.add(mejor)
            coinciden.append({"excel": fila, "chat": disponibles[mejor],
                              "similitud": min(mejor_p, 1.0) if mejor_p < 2 else 1.0,
                              "por_doc": mejor_p >= 2.0})
        else:
            solo_excel.append(fila)

    solo_chat = [it for j, it in enumerate(disponibles) if j not in usados]

    return {
        "coinciden": coinciden,
        "duplicados": duplicados,
        "solo_chat": solo_chat,
        "solo_excel": solo_excel,
        "total_excel": len(filas_excel),
        "total_chat": len(candidatos),
    }


# ------------------------------------------------------------------ reporte

_CAB = Font(bold=True, color="FFFFFF")
_FILL = {
    "ok": PatternFill("solid", fgColor="1E7B34"),
    "dup": PatternFill("solid", fgColor="B58900"),
    "chat": PatternFill("solid", fgColor="C25E00"),
    "excel": PatternFill("solid", fgColor="A32020"),
    "res": PatternFill("solid", fgColor="333333"),
}


COLUMNAS_CARGA = ["#", "GRUPO", "FECHA DE ENVIO", "TRF O DEPOSITO", "TITULAR DE LA CTA",
                  "FECHA TICKET", "HORA TICKET", "CUENTA (CVU)", "MONTO", "Remitente",
                  "CUIL Remitente", "Banco Origen", "Estado", "Origen", "Notas", "Imagen"]


def _fecha_ar(valor) -> str:
    """yyyy-mm-dd -> dd/mm/yyyy. Lo demás se devuelve tal cual."""
    t = str(valor or "").strip()[:10]
    m = re.match(r"(\d{4})-(\d{2})-(\d{2})$", t)
    return f"{m.group(3)}/{m.group(2)}/{m.group(1)}" if m else t


def _hoja_carga(wb, faltantes: list[dict], semana: str = ""):
    """Los faltantes con el formato exacto de la planilla, listos para pegar."""
    ws = wb.create_sheet("📋 Para cargar")
    ws.append(COLUMNAS_CARGA)
    for c in range(1, len(COLUMNAS_CARGA) + 1):
        celda = ws.cell(row=1, column=c)
        celda.font = _CAB
        celda.fill = _FILL["chat"]
        celda.alignment = Alignment(horizontal="center")

    for n, it in enumerate(faltantes, 1):
        perfil = _perfil(it)
        remitente = (it.get("nombre_origen") or it.get("nombre_pie")
                     or it.get("nombre_img") or "")
        doc = it.get("cuit_origen") or it.get("doc_pie") or (sorted(perfil["docs"]) or [""])[0]
        notas = " · ".join(x for x in [
            f"Código de identificación: {it.get('nro_operacion')}" if it.get("nro_operacion") else "",
            f"CVU destinatario: {it.get('cvu_destino')}" if it.get("cvu_destino") else "",
            it.get("detalle") or "",
        ] if x)
        ws.append([
            n,
            it.get("chat", ""),
            semana,
            "TRF",
            it.get("nombre_destino") or "",
            _fecha_ar(it.get("fecha_comp")),
            it.get("hora_comp") or "",
            it.get("cvu_destino") or "",
            _datos_item(it)[1],
            remitente,
            doc,
            it.get("banco") or "",
            "Exitoso" if it.get("estado") == "ok" else (it.get("estado") or ""),
            "grupo",
            notas,
            f'=HYPERLINK("{it.get("link", "")}","Ver imagen")' if it.get("link") else "",
        ])

    anchos = [5, 24, 18, 10, 22, 13, 12, 14, 14, 26, 16, 16, 12, 10, 45, 12]
    for c, a in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(c)].width = a
    ws.freeze_panes = "A2"
    return ws


def _hoja(wb, titulo, cabeceras, filas, color):
    ws = wb.create_sheet(titulo[:31])
    ws.append(cabeceras)
    for c in range(1, len(cabeceras) + 1):
        celda = ws.cell(row=1, column=c)
        celda.font = _CAB
        celda.fill = _FILL[color]
        celda.alignment = Alignment(horizontal="center")
    for f in filas:
        ws.append(f)
    for c in range(1, len(cabeceras) + 1):
        largo = max([len(str(cabeceras[c - 1]))] + [len(str(f[c - 1])) for f in filas[:200] if len(f) >= c] or [10])
        ws.column_dimensions[get_column_letter(c)].width = min(max(largo + 2, 12), 45)
    ws.freeze_panes = "A2"
    return ws


def generar_reporte(res: dict, meta: dict, destino: Path) -> Path:
    wb = Workbook()
    wb.remove(wb.active)

    resumen = wb.create_sheet("Resumen")
    filas_res = [
        ("Chat", meta.get("chat", "")),
        ("Chat ID", meta.get("chat_id", "")),
        ("Excel", meta.get("archivo", "")),
        ("Generado", datetime.now(config.TZ).strftime("%d/%m/%Y %H:%M")),
        ("", ""),
        ("✅ Coinciden", len(res["coinciden"])),
        ("🔁 Duplicados / reenviados", len(res["duplicados"])),
        ("⚠️ En el chat y NO en el Excel", len(res["solo_chat"])),
        ("❌ En el Excel y NO en el chat", len(res["solo_excel"])),
        ("", ""),
        ("Filas del Excel", res["total_excel"]),
        ("Comprobantes del chat", res["total_chat"]),
        ("Monto total coincidente", utils.fmt_monto(
            sum((c["excel"]["monto"] or 0) for c in res["coinciden"]))),
    ]
    for k, v in filas_res:
        resumen.append([k, v])

    if meta.get("grupos"):
        resumen.append(["", ""])
        resumen.append(["Grupos incluidos", ""])
        por_grupo: dict[str, int] = {}
        for s in res["solo_chat"]:
            g = s.get("chat") or str(s.get("chat_id", ""))
            por_grupo[g] = por_grupo.get(g, 0) + 1
        for g in meta["grupos"]:
            resumen.append([f"  {g}", f"{por_grupo.get(g, 0)} sin respaldo en el Excel"])

    for r in range(1, resumen.max_row + 1):
        resumen.cell(row=r, column=1).font = Font(bold=True)
    resumen.column_dimensions["A"].width = 34
    resumen.column_dimensions["B"].width = 40

    _hoja(wb, "✅ Coinciden",
          ["Nombre (Excel)", "Monto (Excel)", "Hoja", "Fila", "Apareo", "Grupo",
           "Nombre (imagen)", "Monto (imagen)", "Similitud", "Fecha msg", "Mensaje"],
          [[c["excel"]["nombre"], c["excel"]["monto"], c["excel"].get("hoja", ""),
            c["excel"]["fila"], "documento/operación" if c.get("por_doc") else "nombre+monto",
            c["chat"].get("chat", ""), _datos_item(c["chat"])[0], _datos_item(c["chat"])[1],
            round(c["similitud"], 2), c["chat"].get("fecha_msg", ""), c["chat"].get("link", "")]
           for c in res["coinciden"]], "ok")

    _hoja(wb, "🔁 Duplicados",
          ["Grupo", "Nombre", "Monto", "Fecha msg", "Remitente", "Original", "Mensaje"],
          [[d.get("chat", ""), _datos_item(d)[0], _datos_item(d)[1], d.get("fecha_msg", ""),
            d.get("remitente", ""), d.get("dup_detalle") or d.get("dup_de", ""), d.get("link", "")]
           for d in res["duplicados"]], "dup")

    _hoja(wb, "⚠️ Solo en chat",
          ["Grupo", "Nombre", "Monto", "Documento", "Estado", "Detalle", "Fecha msg",
           "Remitente", "Mensaje"],
          [[s.get("chat", ""), _datos_item(s)[0], _datos_item(s)[1],
            (sorted(_perfil(s)["docs"]) or [""])[0], s.get("estado", ""),
            s.get("detalle", ""), s.get("fecha_msg", ""), s.get("remitente", ""), s.get("link", "")]
           for s in res["solo_chat"]], "chat")

    _hoja(wb, "❌ Solo en Excel",
          ["Nombre", "Monto", "Documento", "Fecha", "Hora", "Archivo", "Hoja", "Fila"],
          [[s["nombre"], s["monto"], s.get("doc", ""), s.get("fecha", ""), s.get("hora", ""),
            s.get("archivo", ""), s.get("hoja", ""), s["fila"]] for s in res["solo_excel"]],
          "excel")

    if res["solo_chat"]:
        _hoja_carga(wb, res["solo_chat"], meta.get("semana", ""))

    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(destino))
    return destino


def generar_reporte_verificacion(items: list[dict], meta: dict, destino: Path) -> Path:
    """Planilla con TODOS los comprobantes leídos en el chat (sin Excel de por medio)."""
    wb = Workbook()
    wb.remove(wb.active)

    resumen = wb.create_sheet("Resumen")
    conteo: dict[str, int] = {}
    for it in items:
        conteo[it.get("estado", "?")] = conteo.get(it.get("estado", "?"), 0) + 1
    for k, v in [
        ("Chat", meta.get("chat", "")),
        ("Chat ID", meta.get("chat_id", "")),
        ("Desde", meta.get("desde", "todo el historial")),
        ("Hasta", meta.get("hasta", "hoy")),
        ("Mensajes leídos", meta.get("mensajes", "")),
        ("Comprobantes", len(items)),
        ("", ""),
        *[(k2, v2) for k2, v2 in conteo.items()],
    ]:
        resumen.append([k, v])
    resumen.column_dimensions["A"].width = 30
    resumen.column_dimensions["B"].width = 40

    _hoja(wb, "Comprobantes",
          ["Estado", "Nombre (imagen)", "Monto (imagen)", "Nombre (pie)", "Monto (pie)",
           "Detalle", "Banco", "N° operación", "Fecha comprobante", "Fecha msg",
           "Remitente", "Mensaje"],
          [[it.get("estado", ""), it.get("nombre_img", ""), it.get("monto_img"),
            it.get("nombre_pie", ""), it.get("monto_pie"), it.get("detalle", ""),
            it.get("banco", ""), it.get("nro_operacion", ""), it.get("fecha_comp", ""),
            it.get("fecha_msg", ""), it.get("remitente", ""), it.get("link", "")]
           for it in items], "res")

    destino.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(destino))
    return destino


# ------------------------------------------------- importar reporte propio
CABECERAS_VERIF = ("Estado", "Nombre (imagen)", "Monto (imagen)", "Nombre (pie)",
                   "Monto (pie)", "Detalle", "Fecha msg", "Mensaje")


def es_reporte_verificacion(path: str | Path) -> bool:
    """True si el .xlsx es una planilla generada por este bot (no un Excel de cruce)."""
    try:
        wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    except Exception:
        return False
    try:
        if "Comprobantes" not in wb.sheetnames:
            return False
        ws = wb["Comprobantes"]
        fila = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ())
        titulos = {str(c).strip() for c in fila if c}
        return len(titulos & set(CABECERAS_VERIF)) >= 5
    finally:
        wb.close()


def importar_verificacion(path: str | Path) -> tuple[dict, list[dict]]:
    """Reconstruye (meta, items) desde una planilla de verificación del bot."""
    wb = load_workbook(filename=str(path), data_only=True, read_only=True)
    try:
        meta: dict = {}
        if "Resumen" in wb.sheetnames:
            for fila in wb["Resumen"].iter_rows(values_only=True):
                if fila and fila[0]:
                    meta[str(fila[0]).strip()] = fila[1] if len(fila) > 1 else None

        ws = wb["Comprobantes"]
        filas = ws.iter_rows(values_only=True)
        cab = [str(c).strip() if c is not None else "" for c in next(filas, ())]
        idx = {n: i for i, n in enumerate(cab)}

        def val(fila, nombre):
            i = idx.get(nombre)
            if i is None or i >= len(fila):
                return None
            return fila[i]

        try:
            chat_id = int(str(meta.get("Chat ID") or 0))
        except (TypeError, ValueError):
            chat_id = 0
        chat = str(meta.get("Chat") or "")

        items: list[dict] = []
        for fila in filas:
            if not fila or not any(fila):
                continue
            link = str(val(fila, "Mensaje") or "")
            m = re.search(r"/(\d+)\s*$", link)
            msg_id = int(m.group(1)) if m else None
            if not chat_id:
                m2 = re.search(r"/c/(\d+)/", link)
                if m2:
                    chat_id = int(f"-100{m2.group(1)}")

            datos = {
                "monto": val(fila, "Monto (imagen)"),
                "fecha": val(fila, "Fecha comprobante"),
                "nombre_destino": val(fila, "Nombre (imagen)"),
                "nro_operacion": val(fila, "N° operación"),
                "banco": val(fila, "Banco"),
            }
            items.append({
                "chat_id": chat_id,
                "chat": chat,
                "msg_id": msg_id,
                "fecha_msg": str(val(fila, "Fecha msg") or ""),
                "remitente": str(val(fila, "Remitente") or ""),
                "nombre_img": str(val(fila, "Nombre (imagen)") or ""),
                "monto_img": utils.parse_monto(val(fila, "Monto (imagen)")),
                "nombre_pie": str(val(fila, "Nombre (pie)") or ""),
                "monto_pie": utils.parse_monto(val(fila, "Monto (pie)")),
                "banco": str(val(fila, "Banco") or ""),
                "nro_operacion": str(val(fila, "N° operación") or ""),
                "fecha_comp": str(val(fila, "Fecha comprobante") or ""),
                "estado": str(val(fila, "Estado") or "").strip(),
                "detalle": str(val(fila, "Detalle") or ""),
                "link": link,
                "huella": utils.huella_comprobante(datos),
                "importado": True,
            })

        meta_out = {
            "chat_id": chat_id,
            "chat": chat,
            "mensajes": meta.get("Mensajes leídos") or 0,
            "desde": meta.get("Desde"),
            "hasta": meta.get("Hasta"),
        }
        return meta_out, items
    finally:
        wb.close()
